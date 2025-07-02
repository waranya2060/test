from django.core.management.base import BaseCommand, CommandError
from api.models import *
from openpyxl import load_workbook 


class Command(BaseCommand):
    help = 'Closes the specified poll for voting'

    def add_arguments(self, parser):
        #parser.add_argument('poll_ids', nargs='+', type=int)
        pass

    def load(self, wb, sheet_name, column_names):
        print('กำลัง load ... {sheet_name}')
        ws = wb[sheet_name]
        count = int(ws['A2'].value)
        print(f'count = {count}')
        #row4 = [ ws[f'{c}4'].value for c in 'ABCDEFG' ]
        #print( row4 )
        #column_names = ['id', 'rice_type']
        data = []  
        for i in range(count): # 0,1,2,3
            print(f'i = {i}')
            sheet_values = [ ws[f'{chr(65+j)}{4+i}'].value for j in range(len(column_names)) ]
            data.append( dict( (k,v) for k,v in zip(column_names, sheet_values)) )

        return data

    def handle(self, *args, **options):
        # pass
        from openpyxl import load_workbook
        filename = "xlsx/data.xlsx"
        wb = load_workbook(filename, data_only=True)
        #sheets = [ 'Rice_type', 'Work_status', 'Money_status', 'Work' ]

        # print('กำลัง load ... Role')
        # for d in self.load(wb, 'Role', ['id', 'role']):
        #     print(d)
        #     Role(**d).save()

        # print('กำลัง load ... Mechanic_Type')
        # for d in self.load(wb, 'Mechanic_Type', ['id','mechanic_type']):
        #     Mechanic_Type(**d).save()
        
        print('กำลัง load ... Store')
        for d in self.load(wb, 'Store', ['id', 'store_name','store_img', 'store_phone','store_address','store_detail']):
            Store(**d).save()

        # print('กำลัง load ... Product_Type')
        # for d in self.load(wb, 'Product_Type', ['id', 'product_type']):
        #     Product_Type(**d).save()

        # print('กำลัง load ... Product_Status')
        # for d in self.load(wb, 'Product_Status', ['id', 'product_status']):
        #     Product_Status(**d).save()

        print('กำลัง load ... Money_Status')
        for d in self.load(wb, 'Money_Status', ['id', 'money_status']):
            Money_Status(**d).save()
            
        print('กำลัง load ... Delivery_Options')
        for d in self.load(wb, 'Delivery_Options', ['id', 'delivery_options']):
            Delivery_Options(**d).save()

        # print('กำลัง load ... Payment_Options')
        # for d in self.load(wb, 'Payment_Options', ['id', 'payment_options']):
        #     Payment_Options(**d).save()

        # print('กำลัง load ... User')
        # for d in self.load(wb, 'User', ['id', 'username','email', 'password','role']):
        #     # User(**d).save()
        #     role = Role.objects.get(pk=d['role'])
        #     d.pop('role', None)
        #     q = User(**d)
        #     q.role = role
        #     q.save()
        
        # print('กำลัง load ... Customer')
        # for d in self.load(wb, 'Customer', ['id','user','customer_fname', 'customer_lname', 'customer_phone','customer_email','avatar']):
        #     # User(**d).save()
        #     user = User.objects.get(pk=d['user'])
        #     d.pop('user', None)
        #     q = Customer(**d)
        #     q.user = user
        #     q.save()


        # print('กำลัง load ... Mechanic')
        # for d in self.load(wb, 'Mechanic', ['id','user', 'mechanic_fname','mechanic_lname','mechanic_phone', 'mechanic_email','avatar','mechanic_img','mechanic_detail','mechanic_type']):
        #     # Mechanic(**d).save()
        #     user = User.objects.get(pk=d['user'])
        #     d.pop('user', None)
    
        
        #     mechanic_type = Mechanic_Type.objects.get(pk=d['mechanic_type'])
        #     d.pop('mechanic_type', None)
        #     q = Mechanic(**d)
        #     q.user = user
        #     q.mechanic_type = mechanic_type
        #     q.save()
            
            

        

        # print('กำลัง load ... Admin')
        # for d in self.load(wb, 'Admin', ['id', 'user','admin_fname','admin_lname', 'admin_email','avatar']):
        #     # Admin(**d).save()
        #     user = User.objects.get(pk=d['user'])
        #     d.pop('user', None)
        #     q = Admin(**d)
        #     q.user = user
        #     q.save()
            


        # print('กำลัง load ... Product')
        # for d in self.load(wb, 'Product', ['id', 'product_name','product_price', 'product_detail','product_img','product_type','product_status','product_amount']):
        #     # Product(**d).save()
        #     product_type = Product_Type.objects.get(pk=d['product_type'])
        #     d.pop('product_type', None)

        #     product_status = Product_Status.objects.get(pk=d['product_status'])
        #     d.pop('product_status', None)
        #     q = Product(**d)
        #     q.product_type = product_type
        #     q.product_status = product_status
        #     q.save()
            

        

        # print('กำลัง load ... Order')
        # for d in self.load(wb, 'Order', ['id', 'date','user', 'admin','all_price','lat','lng','money_status','delivery_options','payment_options']):
        #     # Order(**d).save()
        #     user = User.objects.get(pk=d['user'])
        #     d.pop('user', None)

        #     admin = Admin.objects.get(pk=d['admin'])
        #     d.pop('admin', None)

        #     money_status = Money_Status.objects.get(pk=d['money_status'])
        #     d.pop('money_status', None)

        #     delivery_options = Delivery_Options.objects.get(pk=d['delivery_options'])
        #     d.pop('delivery_options', None)
 
        #     payment_options = Payment_Options.objects.get(pk=d['payment_options'])
        #     d.pop('payment_options', None)

        #     q = Order(**d)
        #     q.user = user
        #     q.admin = admin
        #     q.money_status = money_status
        #     q.delivery_options = delivery_options
        #     q.payment_options = payment_options
        #     q.save()

        # print('กำลัง load ... Payment')
        # for d in self.load(wb, 'Payment', ['id', 'date','user','order', 'payment_img','payment_options']):
        #     # Payment(**d).save()
        #     user = User.objects.get(pk=d['user'])
        #     d.pop('user', None)
           

        #     order = Order.objects.get(pk=d['order'])
        #     d.pop('order', None)
   
        #     payment_options = Payment_Options.objects.get(pk=d['payment_options'])
        #     d.pop('payment_options', None)
        #     q = Payment(**d)
        #     q.user = user
        #     q.order = order
        #     q.payment_options = payment_options
        #     q.save()

        # print('กำลัง load ... Order_Product')
        # for d in self.load(wb, 'Order_Product', ['id', 'order','product', 'amount']):
        #     # Order_Product(**d).save()
        #     order = Order.objects.get(pk=d['order'])
        #     d.pop('order', None)

        #     product = Product.objects.get(pk=d['product'])
        #     d.pop('product', None)
        #     q = Order_Product(**d)
        #     q.order = order
        #     q.product = product
        #     q.save()

            

        # print('กำลัง load ... Carts')
        # for d in self.load(wb, 'Carts', ['id', 'user','product', 'amount']):
        #     # Carts(**d).save()
        #     user = User.objects.get(pk=d['user'])
        #     d.pop('user', None)

        #     product = Product.objects.get(pk=d['product'])
        #     d.pop('product', None)
        #     q = Carts(**d)
        #     q.user = user
        #     q.product = product
        #     q.save()

        # print('กำลัง load ... Conversations')
        # for d in self.load(wb, 'Conversations', ['id', 'user','message', 'joined_at','updated_at']):
        #     # Conversations(**d).save()
        #     user = User.objects.get(pk=d['user'])
        #     d.pop('user', None)
        #     q = Conversations(**d)
        #     q.user = user
        #     q.save()


        # print('กำลัง load ... Storck')
        # for d in self.load(wb, 'Storck', ['id', 'product','all_products','Sold', 'inventories']):
        #     # Storck(**d).save()
        #     product = Product.objects.get(pk=d['product'])
        #     d.pop('product', None)
        #     q = Storck(**d)
        #     q.product = product
        #     q.save()


      